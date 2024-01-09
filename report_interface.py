# coding:utf-8
import math

from PyQt6.QtSql import QSqlDatabase, QSqlQuery
from PyQt6.QtCore import Qt, QFile, QDateTime, QDate
from PyQt6.QtGui import QPixmap, QColor
from PyQt6.QtWidgets import QFrame, QMessageBox, QFileDialog
from openpyxl.styles import Font, Side, Border, Alignment, PatternFill
from qfluentwidgets import FlowLayout, PushButton, MessageBox, setCustomStyleSheet
from qfluentwidgets import FluentIcon as FIF
from ..UI.DormitoryReport import Ui_Frame
from openpyxl import Workbook


class ReportInterface(Ui_Frame, QFrame):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setupUi(self)
        self.setObjectName('reportinterface')
        self.scrollAreaWidgetContents.parent().parent().setStyleSheet("background-color: rgba(0,0,0,0);border: none;")

        self.db = QSqlDatabase.addDatabase('QSQLITE')
        self.db.setDatabaseName('./xldb.db')
        if not self.db.open():
            QMessageBox.critical(self, 'Database Connection', self.db.lastError().text())

        self.flowLayout = FlowLayout(self.SimpleCardWidget_2, needAni=False)

        studentForm = PushButton(self)
        studentForm.setText('导出学生列表')
        studentForm.clicked.connect(self.__onStudentForm)

        bedListForm = PushButton(self)
        bedListForm.setText('导出登记名单')
        bedListForm.clicked.connect(self.__onRosterForm)

        lookupTable = PushButton(self)
        lookupTable.setText('导出查宿表')
        lookupTable.clicked.connect(self.__onLookupTable)

        self.flowLayout.addWidget(studentForm)
        self.flowLayout.addWidget(bedListForm)
        self.flowLayout.addWidget(lookupTable)

        self.__inteWindow()

    def __onStudentForm(self):
        folder = QFileDialog.getSaveFileName(self, "文件保存", f"./app/download/学生列表-{QDateTime.currentDateTime().toString('yyyy-MM-dd-HH-mm-ss')}", "Excel工作簿 (*.xlsx)")
        if folder[0] != '':
            query = QSqlQuery(self.db)
            wb = Workbook()
            ws = wb.active

            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 25
            ws.column_dimensions['K'].width = 15

            ws.merge_cells('A1:K1')
            ws.append(["序号", "学生姓名", "学生电话", "家长姓名", "家长电话", "性别", "班级", "宿舍号", "床位", "地址", "入住日期"])

            if query.exec("SELECT * FROM students"):
                while query.next():
                    ws.append([query.value("id"),
                               query.value("name"),
                               query.value("phone"),
                               query.value("parent_name"),
                               query.value("parent_phone"),
                               query.value("gender"),
                               query.value("class"),
                               query.value("dormitory_id"),
                               query.value("bed_number"),
                               query.value("address"),
                               query.value("arrival_date")])

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name='Calibri', size=14, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws['A1'] = '花名册'
            ws['A1'].font = Font(name='Calibri', size=28, bold=True)
            ws.row_dimensions[1].height = 50

            try:
                wb.save(folder[0])
            except:
                while True:
                    w = MessageBox('文件无法处理', '请关闭文件后重试', self)
                    w.yesButton.setText('重试')
                    if w.exec():
                        try:
                            wb.save(folder[0])
                            break
                        except:
                            w.exec()
                    else:
                        break

    def __onRosterForm(self):
        folder = QFileDialog.getSaveFileName(self, "文件保存", f"./app/download/登记名单-{QDateTime.currentDateTime().toString('yyyy-MM-dd-HH-mm-ss')}", "Excel工作簿 (*.xlsx)")
        if folder[0] != '':
            query = QSqlQuery(self.db)
            wb = Workbook()
            wb.remove(wb['Sheet'])
            students_list = []
            if query.exec("SELECT id, name, gender, class, dormitory_id FROM students"):
                while query.next():
                    students_list.append([query.value("id"),
                                          query.value("name"),
                                          query.value("gender"),
                                          query.value("class"),
                                          query.value("dormitory_id")])

            for i in range(math.ceil(len(students_list)/70)):
                ws = wb.create_sheet(f'第{i+1}页')
                wb.active = ws
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 8
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 8
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 8
                ws.column_dimensions['I'].width = 12
                ws.column_dimensions['J'].width = 12
                ws.merge_cells('A1:J1')
                ws.append(["序号", "学生姓名", "性别", "班级", "宿舍号",
                           "序号", "学生姓名", "性别", "班级", "宿舍号"])
                if len(students_list) > 1:
                    data = students_list[i*70:(i+1)*70]
                    mid_point = len(data) // 2
                    first_half = data[:mid_point]
                    second_half = data[mid_point:]
                    new_list = [x for pair in zip(first_half, second_half) for x in pair]
                    for sublist in [new_list[i:i + 2] for i in range(0, len(new_list), 2)]:
                        if len(sublist) == 2:
                            ws.append(sublist[0] + sublist[1])
                        else:
                            ws.append(sublist[0])
                else:
                    ws.append(students_list[0])

                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = Font(name='Calibri', size=14, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                ws['A1'] = '内宿生登记名单'
                ws['A1'].font = Font(name='Calibri', size=28, bold=True)
                ws['A1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['B1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['C1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['D1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['E1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['F1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['G1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['H1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['I1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws['J1'].border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
                ws.row_dimensions[1].height = 50
            if len(students_list) != 0:
                wb.active = wb['第1页']
                try:
                    wb.save(folder[0])
                except:
                    while True:
                        w = MessageBox('文件无法处理', '请关闭文件后重试', self)
                        w.yesButton.setText('重试')
                        if w.exec():
                            try:
                                wb.save(folder[0])
                                break
                            except:
                                w.exec()
                        else:
                            break
            else:
                w = MessageBox('无数据', '请添加数据后再导出', self)
                qss = '''
                                        PushButton{background-color: #E53935; border-color: #E53935; color: #FFFFFF}
                                        PushButton::hover{background-color: #EF5350; border-color: #EF5350}
                                        '''
                setCustomStyleSheet(w.yesButton, qss, qss)
                w.cancelButton.hide()
                if w.exec():
                    pass

    def __onLookupTable(self):
        folder = QFileDialog.getSaveFileName(self, "文件保存", f"./app/download/查宿表-{QDateTime.currentDateTime().toString('yyyy-MM-dd-HH-mm-ss')}", "Excel工作簿 (*.xlsx)")
        if folder[0] != '':
            query = QSqlQuery(self.db)
            wb = Workbook()
            wb.remove(wb['Sheet'])

            year = QDate.currentDate().toString("yyyy")
            moon = QDate.currentDate().toString("MM")
            entry = []

            if query.exec("SELECT dormitory_name, bed_number FROM dormitories"):
                queryNext = query.next()
                while queryNext:
                    entry.append(queryNext)
                    ws = wb.create_sheet(f'{query.value(0)}宿舍')
                    wb.active = ws

                    ws.column_dimensions['A'].width = 4
                    ws.column_dimensions['B'].width = 10
                    ws.column_dimensions['C'].width = 10
                    ws.column_dimensions['D'].width = 16
                    ws.column_dimensions['E'].width = 16
                    ws.column_dimensions['F'].width = 3
                    ws.column_dimensions['G'].width = 3
                    ws.column_dimensions['H'].width = 3
                    ws.column_dimensions['I'].width = 3
                    ws.column_dimensions['J'].width = 3
                    ws.column_dimensions['K'].width = 3
                    ws.column_dimensions['L'].width = 3
                    ws.column_dimensions['M'].width = 3
                    ws.column_dimensions['N'].width = 3
                    ws.column_dimensions['O'].width = 3
                    ws.column_dimensions['P'].width = 3
                    ws.column_dimensions['Q'].width = 3
                    ws.column_dimensions['R'].width = 3
                    ws.column_dimensions['S'].width = 3
                    ws.column_dimensions['T'].width = 3
                    ws.column_dimensions['U'].width = 3
                    ws.column_dimensions['V'].width = 3
                    ws.column_dimensions['W'].width = 3
                    ws.column_dimensions['X'].width = 3
                    ws.column_dimensions['Y'].width = 3
                    ws.column_dimensions['Z'].width = 3
                    ws.column_dimensions['AA'].width = 3
                    ws.column_dimensions['AB'].width = 3
                    ws.column_dimensions['AC'].width = 3
                    ws.merge_cells('A1:AC1')

                    if int(moon) < 6:
                        ws['A1'] = f'{year}春季({query.value(0)})宿舍查宿登记表)'
                    else:
                        ws['A1'] = f'{year}秋季({query.value(0)})宿舍查宿登记表)'

                    ws.append(['序号', '姓名', '年级', '学生电话', '家长电话',
                               '日', '一', '二', '三', '四', '五',
                               '日', '一', '二', '三', '四', '五',
                               '日', '一', '二', '三', '四', '五',
                               '日', '一', '二', '三', '四', '五'])

                    lookupQuery = QSqlQuery(self.db)
                    if lookupQuery.exec(f"SELECT name, class, phone, parent_phone FROM students WHERE dormitory_id = '{query.value(0)}'"):
                        nub = 0
                        while lookupQuery.next():
                            nub += 1
                            ws.append([nub,
                                       lookupQuery.value("name"),
                                       lookupQuery.value("class"),
                                       lookupQuery.value("phone"),
                                       lookupQuery.value("parent_phone")])
                        ws.merge_cells(f'A{nub+3}:AC{nub+3}')
                        ws[f'A{nub+3}'] = '正常√   晚归○   未到×   请假Ø                 负责人签名:'

                    for row in ws.iter_rows():
                        for cell in row:
                            cell.font = Font(name='Calibri', size=12, bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    ws['A1'].font = Font(name='Calibri', size=28, bold=True)
                    queryNext = query.next()

            if len(entry) > 0:
                try:
                    wb.save(folder[0])
                except:
                    while True:
                        w = MessageBox('文件无法处理', '请关闭文件后重试', self)
                        w.yesButton.setText('重试')
                        if w.exec():
                            try:
                                wb.save(folder[0])
                                break
                            except:
                                w.exec()
                        else:
                            break
            else:
                w = MessageBox('无数据', '请添加数据后再导出', self)
                qss = '''
                        PushButton{background-color: #E53935; border-color: #E53935; color: #FFFFFF}
                        PushButton::hover{background-color: #EF5350; border-color: #EF5350}
                        '''
                setCustomStyleSheet(w.yesButton, qss, qss)
                w.cancelButton.hide()
                if w.exec():
                    pass

    def __inteWindow(self):
        self.PixmapLabel_Room.setPixmap(QPixmap(':/gallery/images/Room.png').scaled(
            50, 50, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        self.PixmapLabel_Bed.setPixmap(QPixmap(':/gallery/images/Bed.png').scaled(
            50, 50, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        self.PixmapLabel_Personnel.setPixmap(QPixmap(':/gallery/images/Student.png').scaled(
            50, 50, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        self.PixmapLabel_Vacant.setPixmap(QPixmap(':/gallery/images/None.png').scaled(
            50, 50, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))

        dcolor = QColor(127, 127, 127)
        self.TitleLabel_Room.setTextColor(dcolor, dcolor)
        self.TitleLabel_Bed.setTextColor(dcolor, dcolor)
        self.TitleLabel_Personnel.setTextColor(dcolor, dcolor)
        self.TitleLabel_Vacant.setTextColor(dcolor, dcolor)

        self.updateReport()
        self.loadText()

    def updateReport(self):
        query = QSqlQuery(self.db)

        if query.exec("SELECT COUNT(*) FROM students;"):
            query.next()
            self.TitleLabel_PersonnelNub.setText(str(query.value(0)))

        if query.exec("SELECT COUNT(*) FROM dormitories;"):
            query.next()
            self.TitleLabel_RoomNub.setText(str(query.value(0)))

        if query.exec("SELECT SUM(bed_number) FROM dormitories;"):
            query.next()
            if query.value(0) != '':
                self.TitleLabel_BedNub.setText(str(query.value(0)))
            else:
                self.TitleLabel_BedNub.setText('0')

        if query.exec("SELECT COUNT(*) FROM students WHERE dormitory_id IS NULL OR dormitory_id = ''"):
            query.next()
            Occupancy = int(self.TitleLabel_PersonnelNub.text()) - int(query.value(0))
            if self.TitleLabel_BedNub.text() != '' and int(self.TitleLabel_BedNub.text()) > Occupancy:
                self.TitleLabel_VacantNub.setText(str(int(self.TitleLabel_BedNub.text()) - Occupancy))
            else:
                self.TitleLabel_VacantNub.setText('0')

    def loadText(self):
        self.updateReport()
        file = QFile('oh.log')
        if file.open(QFile.OpenModeFlag.ReadOnly | QFile.OpenModeFlag.Text):
            content = file.readAll()
            self.StrongBodyLabel.setText(str(content, 'gbk'))
        else:
            print('Failed to open file')
